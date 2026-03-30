#!/usr/bin/env python3
import argparse
import base64
import csv
import json
import os
import sys
import time
import uuid
from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional, Tuple
from urllib.parse import quote

try:
    import requests
except ImportError:
    print("ERROR: requests is not installed. Run: pip install requests", file=sys.stderr)
    sys.exit(1)


def lazy_import_openpyxl():
    try:
        import openpyxl  # type: ignore
        return openpyxl
    except ImportError:
        print("ERROR: openpyxl is not installed. Run: pip install openpyxl", file=sys.stderr)
        sys.exit(1)


TOKEN_URL = "https://marketplace.walmartapis.com/v3/token"
DELETE_URL_TMPL = "https://marketplace.walmartapis.com/v3/items/{sku}"
DEFAULT_WM_SVC_NAME = os.getenv("WALMART_WM_SVC_NAME", "Walmart Marketplace")
TERMINAL_SKIP_OUTCOMES = {"deleted", "not_found"}
TRANSIENT = {408, 425, 429, 500, 502, 503, 504, 520, 521, 522, 524}
KNOWN_HEADERS = {"sku", "item", "item id", "itemid", "seller sku"}


@dataclass
class DeleteResult:
    sku: str
    status_code: int
    outcome: str
    detail: str
    attempts: int


def correlation_id() -> str:
    return str(uuid.uuid4())


class RateLimiter:
    def __init__(self, per_minute: int) -> None:
        self.interval = 60.0 / max(1, per_minute)
        self.last = 0.0

    def wait(self) -> None:
        now = time.monotonic()
        delta = now - self.last
        if delta < self.interval:
            time.sleep(self.interval - delta)
        self.last = time.monotonic()


def get_token(client_id: str, client_secret: str, timeout: int = 60) -> str:
    pair = f"{client_id}:{client_secret}".encode("ascii")
    basic = base64.b64encode(pair).decode("ascii")
    headers = {
        "Authorization": f"Basic {basic}",
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
        "WM_SVC.NAME": DEFAULT_WM_SVC_NAME,
        "WM_QOS.CORRELATION_ID": correlation_id(),
    }
    channel_type = os.getenv("WALMART_CHANNEL_TYPE")
    if channel_type:
        headers["WM_CONSUMER.CHANNEL.TYPE"] = channel_type

    resp = requests.post(
        TOKEN_URL,
        headers=headers,
        data="grant_type=client_credentials",
        timeout=timeout,
    )
    if resp.status_code != 200:
        raise RuntimeError(f"Token request failed ({resp.status_code}): {resp.text[:1000]}")
    data = resp.json()
    token = data.get("access_token")
    if not token:
        raise RuntimeError(f"Token response missing access_token: {data}")
    return token


def build_api_headers(token: str) -> dict:
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "WM_SEC.ACCESS_TOKEN": token,
        "WM_SVC.NAME": DEFAULT_WM_SVC_NAME,
        "WM_QOS.CORRELATION_ID": correlation_id(),
    }
    channel_type = os.getenv("WALMART_CHANNEL_TYPE")
    if channel_type:
        headers["WM_CONSUMER.CHANNEL.TYPE"] = channel_type
    return headers


def delete_one_sku(
    session: requests.Session,
    token: str,
    sku: str,
    max_attempts: int = 5,
    timeout: int = 60,
) -> DeleteResult:
    encoded_sku = quote(sku, safe="")
    url = DELETE_URL_TMPL.format(sku=encoded_sku)

    last_status = 0
    last_text = ""
    for attempt in range(1, max_attempts + 1):
        headers = build_api_headers(token)
        try:
            resp = session.delete(url, headers=headers, timeout=timeout)
        except KeyboardInterrupt:
            raise
        except requests.RequestException as e:
            last_status = 0
            last_text = str(e)
            if attempt < max_attempts:
                time.sleep(min(2 ** (attempt - 1), 8) + 0.25)
                continue
            return DeleteResult(sku, 0, "network_error", last_text, attempt)

        last_status = resp.status_code
        last_text = resp.text[:1000]

        if resp.status_code in (200, 204):
            return DeleteResult(sku, resp.status_code, "deleted", last_text, attempt)
        if resp.status_code == 404:
            return DeleteResult(sku, resp.status_code, "not_found", last_text, attempt)
        if resp.status_code == 400:
            return DeleteResult(sku, resp.status_code, "bad_request", last_text, attempt)
        if resp.status_code == 401:
            return DeleteResult(sku, resp.status_code, "unauthorized", last_text, attempt)
        if resp.status_code == 403:
            return DeleteResult(sku, resp.status_code, "forbidden", last_text, attempt)
        if resp.status_code == 409:
            return DeleteResult(sku, resp.status_code, "conflict", last_text, attempt)
        if resp.status_code in TRANSIENT and attempt < max_attempts:
            time.sleep(min(2 ** (attempt - 1), 8) + 0.25)
            continue
        return DeleteResult(sku, resp.status_code, "failed", last_text, attempt)

    return DeleteResult(sku, last_status, "failed", last_text, max_attempts)


def normalize_sku(value) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    if s.endswith(".0") or s.endswith(".00"):
        head = s.split(".", 1)[0]
        if head.isdigit():
            s = head
    return s


def dedupe_preserve_order(items: Iterable[str]) -> List[str]:
    seen = set()
    out: List[str] = []
    for item in items:
        if item not in seen:
            seen.add(item)
            out.append(item)
    return out


def read_skus_from_csv(path: str, column: Optional[str]) -> List[str]:
    skus: List[str] = []
    with open(path, "r", newline="", encoding="utf-8-sig") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(f, dialect)
        rows = list(reader)

    if not rows:
        return []

    col_idx: Optional[int] = None
    if column:
        if column.isdigit():
            col_idx = int(column)
        else:
            headers = [h.strip().lower() for h in rows[0]]
            if column.lower() in headers:
                col_idx = headers.index(column.lower())
            elif len(column) == 1 and column.isalpha():
                col_idx = ord(column.upper()) - ord("A")

    start_row = 0
    if col_idx is None:
        headers = [h.strip().lower() for h in rows[0]]
        for i, h in enumerate(headers):
            if h in KNOWN_HEADERS:
                col_idx = i
                start_row = 1
                break
    if col_idx is None:
        col_idx = 0

    for row in rows[start_row:]:
        if col_idx < len(row):
            sku = normalize_sku(row[col_idx])
            if sku:
                skus.append(sku)
    return dedupe_preserve_order(skus)


def read_skus_from_xlsx(path: str, sheet_name: Optional[str], column: Optional[str]) -> List[str]:
    openpyxl = lazy_import_openpyxl()
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    col_idx: Optional[int] = None
    start_row = 1

    first_row = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

    if column:
        if column.isdigit():
            col_idx = int(column)
        elif len(column) == 1 and column.isalpha():
            col_idx = ord(column.upper()) - ord("A") + 1
        else:
            normalized = [str(x).strip().lower() if x is not None else "" for x in first_row]
            if column.lower() in normalized:
                col_idx = normalized.index(column.lower()) + 1
                start_row = 2

    if col_idx is None:
        normalized = [str(x).strip().lower() if x is not None else "" for x in first_row]
        for i, h in enumerate(normalized, start=1):
            if h in KNOWN_HEADERS:
                col_idx = i
                start_row = 2
                break
    if col_idx is None:
        col_idx = 1
        first = normalize_sku(first_row[0]) if first_row else None
        if first and first.lower() in KNOWN_HEADERS:
            start_row = 2

    skus: List[str] = []
    for row in ws.iter_rows(min_row=start_row, min_col=col_idx, max_col=col_idx, values_only=True):
        sku = normalize_sku(row[0])
        if sku:
            skus.append(sku)

    wb.close()
    return dedupe_preserve_order(skus)


def read_skus(path: str, sheet_name: Optional[str], column: Optional[str]) -> List[str]:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return read_skus_from_csv(path, column)
    if ext in {".xlsx", ".xlsm"}:
        return read_skus_from_xlsx(path, sheet_name, column)
    raise RuntimeError(f"Unsupported input file type: {ext}")


def ensure_log_header(path: str) -> None:
    if os.path.exists(path) and os.path.getsize(path) > 0:
        return
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["sku", "status_code", "outcome", "attempts", "detail"])


def append_log_row(path: str, row: DeleteResult) -> None:
    ensure_log_header(path)
    with open(path, "a", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([row.sku, row.status_code, row.outcome, row.attempts, row.detail])
        f.flush()


def load_existing_log(path: str) -> Tuple[List[DeleteResult], Dict[str, DeleteResult]]:
    if not os.path.exists(path) or os.path.getsize(path) == 0:
        return [], {}

    rows: List[DeleteResult] = []
    latest: Dict[str, DeleteResult] = {}
    with open(path, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for r in reader:
            try:
                item = DeleteResult(
                    sku=(r.get("sku") or "").strip(),
                    status_code=int(r.get("status_code") or 0),
                    outcome=(r.get("outcome") or "").strip(),
                    attempts=int(r.get("attempts") or 0),
                    detail=r.get("detail") or "",
                )
            except Exception:
                continue
            if not item.sku:
                continue
            rows.append(item)
            latest[item.sku] = item
    return rows, latest


def write_summary(path: str, last_results: Dict[str, DeleteResult], meta: Optional[dict] = None) -> None:
    totals: Dict[str, int] = {}
    for r in last_results.values():
        totals[r.outcome] = totals.get(r.outcome, 0) + 1

    lines = [
        "Walmart delete summary",
        "======================",
        f"Unique SKUs with log entries: {len(last_results)}",
    ]
    if meta:
        for key in [
            "input_file",
            "total_extracted",
            "resumed_from_index",
            "remaining_to_process_at_start",
            "auto_skipped_from_prior_log",
            "checkpoint_file",
        ]:
            if key in meta:
                lines.append(f"{key}: {meta[key]}")
    for k in sorted(totals):
        lines.append(f"{k}: {totals[k]}")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")


def save_checkpoint(path: str, payload: dict) -> None:
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)
    os.replace(tmp, path)


def load_checkpoint(path: str) -> Optional[dict]:
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def remove_checkpoint(path: str) -> None:
    try:
        if os.path.exists(path):
            os.remove(path)
    except OSError:
        pass


def main() -> int:
    ap = argparse.ArgumentParser(description="Delete Walmart SKUs directly from a CSV/XLSX source with resume + token refresh.")
    ap.add_argument("--in-file", required=True, help="Path to CSV or XLSX containing SKUs.")
    ap.add_argument("--sheet", help="Worksheet name for XLSX input.")
    ap.add_argument("--column", help="Column name, letter, or zero/one-based index depending on file type. Example: sku or D")
    ap.add_argument("--log-csv", default="walmart_delete_results.csv", help="Where to save per-SKU results.")
    ap.add_argument("--summary-txt", default="walmart_delete_summary.txt", help="Where to save summary text.")
    ap.add_argument("--checkpoint-file", default="walmart_delete_checkpoint.json", help="Resume checkpoint file. Saved after every SKU.")
    ap.add_argument("--rate-per-minute", type=int, default=300, help="Delete rate. Default is a gentler 300/min.")
    ap.add_argument("--start-at", type=int, default=1, help="1-based starting row within extracted SKU list, for manual resuming.")
    ap.add_argument("--max-count", type=int, help="Maximum number of extracted SKUs to process.")
    ap.add_argument("--dry-run", action="store_true", help="Read and log SKUs without deleting them.")
    ap.add_argument("--fresh-run", action="store_true", help="Ignore existing log/checkpoint and start from scratch.")
    ap.add_argument("--refresh-token-every-minutes", type=float, default=10.0, help="Proactively refresh the access token every N minutes during long runs. Set 0 to disable.")
    ap.add_argument("--stop-on-persistent-401", action="store_true", default=True, help="Stop and save checkpoint if a SKU still returns 401 after token refresh.")
    args = ap.parse_args()

    client_id = os.getenv("WALMART_CLIENT_ID")
    client_secret = os.getenv("WALMART_CLIENT_SECRET")
    if not client_id or not client_secret:
        print("ERROR: Set WALMART_CLIENT_ID and WALMART_CLIENT_SECRET in your environment first.", file=sys.stderr)
        return 2

    try:
        all_skus = read_skus(args.in_file, args.sheet, args.column)
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 2

    if not all_skus:
        print("ERROR: No SKUs found in input file.", file=sys.stderr)
        return 2

    start_index = max(1, args.start_at)
    previous_rows: List[DeleteResult] = []
    latest_results: Dict[str, DeleteResult] = {}
    if not args.fresh_run:
        previous_rows, latest_results = load_existing_log(args.log_csv)

    checkpoint = None if args.fresh_run else load_checkpoint(args.checkpoint_file)
    if checkpoint:
        cp_input = checkpoint.get("input_file")
        cp_sheet = checkpoint.get("sheet")
        cp_col = checkpoint.get("column")
        if cp_input == os.path.abspath(args.in_file) and cp_sheet == args.sheet and cp_col == args.column:
            cp_next = int(checkpoint.get("next_source_index", 1) or 1)
            if cp_next > start_index:
                start_index = cp_next

    if args.max_count is not None:
        capped_end = start_index - 1 + args.max_count
    else:
        capped_end = len(all_skus)

    pre_done = {sku for sku, r in latest_results.items() if r.outcome in TERMINAL_SKIP_OUTCOMES}
    total_extracted = len(all_skus)

    to_process: List[Tuple[int, str]] = []
    auto_skipped = 0
    for source_index, sku in enumerate(all_skus, start=1):
        if source_index < start_index or source_index > capped_end:
            continue
        if sku in pre_done:
            auto_skipped += 1
            continue
        to_process.append((source_index, sku))

    print(f"Loaded {total_extracted} unique SKU(s) from {args.in_file}")
    if previous_rows and not args.fresh_run:
        print(f"Found existing log with {len(latest_results)} unique prior SKU result(s); auto-skipping {auto_skipped} already deleted/not_found SKU(s).")
    if checkpoint and not args.fresh_run:
        print(f"Resuming from checkpoint/source index {start_index}.")
    if args.max_count is not None:
        print(f"Max count applied: {args.max_count}")
    print(f"Remaining SKU(s) to process now: {len(to_process)}")

    meta = {
        "input_file": os.path.abspath(args.in_file),
        "total_extracted": total_extracted,
        "resumed_from_index": start_index,
        "remaining_to_process_at_start": len(to_process),
        "auto_skipped_from_prior_log": auto_skipped,
        "checkpoint_file": os.path.abspath(args.checkpoint_file),
    }

    if args.dry_run:
        for _, sku in to_process:
            latest_results[sku] = DeleteResult(sku, 0, "dry_run", "Not deleted; dry run only.", 0)
        write_summary(args.summary_txt, latest_results, meta)
        print(f"Dry run complete. Wrote {args.summary_txt}")
        return 0

    try:
        token = get_token(client_id, client_secret)
        token_acquired_at = time.monotonic()
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 2

    limiter = RateLimiter(args.rate_per_minute)
    session = requests.Session()
    newly_logged = 0

    def maybe_refresh_token(force: bool = False) -> None:
        nonlocal token, token_acquired_at
        if force:
            token = get_token(client_id, client_secret)
            token_acquired_at = time.monotonic()
            return
        if args.refresh_token_every_minutes and args.refresh_token_every_minutes > 0:
            age_minutes = (time.monotonic() - token_acquired_at) / 60.0
            if age_minutes >= args.refresh_token_every_minutes:
                print(f"Refreshing access token after {age_minutes:.1f} minute(s)...")
                token = get_token(client_id, client_secret)
                token_acquired_at = time.monotonic()

    try:
        for run_pos, (source_index, sku) in enumerate(to_process, start=1):
            maybe_refresh_token(force=False)
            limiter.wait()
            result = delete_one_sku(session, token, sku)

            if result.status_code == 401:
                print(f"[{run_pos}/{len(to_process)} | source {source_index}/{total_extracted}] {sku} -> unauthorized (401); refreshing token and retrying same SKU...")
                try:
                    token = get_token(client_id, client_secret)
                    token_acquired_at = time.monotonic()
                except Exception as e:
                    append_log_row(args.log_csv, result)
                    latest_results[sku] = result
                    newly_logged += 1
                    save_checkpoint(args.checkpoint_file, {
                        **meta,
                        "sheet": args.sheet,
                        "column": args.column,
                        "next_source_index": source_index,
                        "last_sku": sku,
                        "last_outcome": "unauthorized_token_refresh_failed",
                        "last_status_code": 401,
                    })
                    write_summary(args.summary_txt, latest_results, meta)
                    print(f"ERROR: Token refresh failed after 401: {e}", file=sys.stderr)
                    return 1

                limiter.wait()
                result = delete_one_sku(session, token, sku)

            latest_results[sku] = result
            append_log_row(args.log_csv, result)
            newly_logged += 1

            print(f"[{run_pos}/{len(to_process)} | source {source_index}/{total_extracted}] {sku} -> {result.outcome} ({result.status_code})")

            if result.status_code == 401 and args.stop_on_persistent_401:
                save_checkpoint(args.checkpoint_file, {
                    **meta,
                    "sheet": args.sheet,
                    "column": args.column,
                    "next_source_index": source_index,
                    "last_sku": sku,
                    "last_outcome": result.outcome,
                    "last_status_code": result.status_code,
                })
                write_summary(args.summary_txt, latest_results, meta)
                print(f"Stopped after persistent 401 on SKU {sku}. Re-run the same command to resume from that SKU.", file=sys.stderr)
                return 1

            save_checkpoint(args.checkpoint_file, {
                **meta,
                "sheet": args.sheet,
                "column": args.column,
                "next_source_index": source_index + 1,
                "last_sku": sku,
                "last_outcome": result.outcome,
                "last_status_code": result.status_code,
            })

            if newly_logged % 50 == 0:
                write_summary(args.summary_txt, latest_results, meta)

    except KeyboardInterrupt:
        write_summary(args.summary_txt, latest_results, meta)
        # Save checkpoint to the current position if possible. If nothing processed yet in this run,
        # keep the prior/start position so the next run retries safely from there.
        retry_from = start_index
        if checkpoint and checkpoint.get("next_source_index"):
            retry_from = int(checkpoint.get("next_source_index") or start_index)
        latest_cp = load_checkpoint(args.checkpoint_file)
        if latest_cp and latest_cp.get("next_source_index"):
            retry_from = int(latest_cp.get("next_source_index") or retry_from)
        save_checkpoint(args.checkpoint_file, {
            **meta,
            "sheet": args.sheet,
            "column": args.column,
            "next_source_index": retry_from,
            "last_outcome": "keyboard_interrupt",
        })
        print("Interrupted. Progress has been saved. Re-run the same command to continue.", file=sys.stderr)
        return 130

    write_summary(args.summary_txt, latest_results, meta)
    remove_checkpoint(args.checkpoint_file)
    print(f"Done. Results: {args.log_csv}")
    print(f"Done. Summary: {args.summary_txt}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
